"""
File Handler for Student Management App
Handles file upload, parsing, and text extraction
"""

import os
import tempfile
import pandas as pd
import logging
import re
import unicodedata
from pathlib import Path

logger = logging.getLogger(__name__)

# Try to import optional dependencies
try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    logger.warning("python-docx not installed. DOCX file support disabled.")

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not installed. XLSX file support disabled.")

def secure_filename(filename):
    """Fallback secure_filename implementation"""
    filename = filename.replace(' ', '_')
    return re.sub(r'[^a-zA-Z0-9._-]', '', filename)

def extract_text_from_file(file_path, file_extension):
    """Extract text from various file formats"""
    try:
        if file_extension in ['.txt', '.csv']:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif file_extension == '.docx' and DOCX_AVAILABLE:
            doc = docx.Document(file_path)
            return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
        elif file_extension in ['.xlsx', '.xls'] and XLSX_AVAILABLE:
            # Try pandas first for better data handling
            try:
                df = pd.read_excel(file_path)
                return df.to_string(index=False)
            except:
                # Fallback to openpyxl
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                text_content = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        row_text = [str(cell) if cell is not None else '' for cell in row]
                        text_content.append('\t'.join(row_text))
                return '\n'.join(text_content)
        else:
            logger.warning(f"Unsupported file extension: {file_extension}")
            return ""
    except Exception as e:
        logger.error(f"Error extracting text from {file_path}: {e}")
        return ""

def is_vietnamese_name(line, words):
    """Check if line contains a Vietnamese name"""
    if (len(words) >= 2 and len(words) <= 5 and 
        not any(char.isdigit() for char in line) and 
        5 <= len(line) <= 50 and
        all(len(word) >= 1 for word in words)):
        
        # Check for Vietnamese characteristics
        has_vietnamese_chars = any(char in 'àáâãèéêìíòóôõùúăđĩũơưăạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵýỷỹ' for char in line.lower())
        has_proper_caps = all(word[0].isupper() and (len(word) == 1 or word[1:].islower()) for word in words)
        is_all_caps = all(word.isupper() for word in words)
        has_common_vn_chars = any(char in 'ăôưđ' for char in line.lower())
        
        # Exclude technical terms
        tech_words = {'file', 'data', 'csv', 'excel', 'document', 'text', 'import', 'export', 'column', 'row'}
        has_tech_words = any(word.lower() in tech_words for word in words)
        
        return (has_vietnamese_chars or has_proper_caps or is_all_caps or has_common_vn_chars) and not has_tech_words
    
    return False

def is_class_pattern(line):
    """Check if line contains a class pattern"""
    # Common Vietnamese class patterns
    class_patterns = [
        r'^\d{1,2}[A-Z]{1,4}\d{0,2}$',  # e.g., 14TCLC3, 15A1
        r'^[A-Z]{1,4}\d{1,3}$',         # e.g., CNTT1, IT01
        r'^Lớp\s+\d{1,2}[A-Z]{1,4}\d{0,2}$',  # e.g., Lớp 14TCLC3
        r'^\d{1,2}[A-Z][a-z]*\d{0,2}$', # e.g., 14Tin1
    ]
    
    for pattern in class_patterns:
        if re.match(pattern, line, re.IGNORECASE):
            return True
    
    return False

def is_date_pattern(line):
    """Check if line contains a date pattern"""
    # Common date patterns
    date_patterns = [
        r'\d{1,2}/\d{1,2}/\d{4}',      # DD/MM/YYYY
        r'\d{1,2}-\d{1,2}-\d{4}',      # DD-MM-YYYY
        r'\d{4}/\d{1,2}/\d{1,2}',      # YYYY/MM/DD
        r'\d{4}-\d{1,2}-\d{1,2}',      # YYYY-MM-DD
    ]
    
    for pattern in date_patterns:
        if re.search(pattern, line):
            return True
    
    return False

def parse_mixed_content(line):
    """Parse mixed content line for multiple data types"""
    results = {
        'mssv': [],
        'names': [],
        'classes': [],
        'dates': []
    }
    
    # Extract MSSV patterns
    mssv_patterns = [
        r'\b(20\d{8})\b',      # 10-digit starting with 20
        r'\b(20\d{6})\b',      # 8-digit starting with 20
        r'\b(\d{6,10})\b'      # General 6-10 digit pattern
    ]
    
    for pattern in mssv_patterns:
        matches = re.findall(pattern, line)
        for match in matches:
            if 6 <= len(match) <= 10:
                results['mssv'].append(match)
    
    # Extract Vietnamese name patterns
    name_pattern = r'\b([A-ZÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠƯĂẠẢẤẦẨẪẬẮẰẲẴẶẸẺẼẾỀỂỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪỬỮỰỲỴÝỶỸ][a-zàáâãèéêìíòóôõùúăđĩũơưăạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵýỷỹ]+(?:\s+[A-ZÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠƯĂẠẢẤẦẨẪẬẮẰẲẴẶẸẺẼẾỀỂỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪỬỮỰỲỴÝỶỸ][a-zàáâãèéêìíòóôõùúăđĩũơưăạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵýỷỹ]+)+)\b'
    name_matches = re.findall(name_pattern, line)
    for match in name_matches:
        words = match.split()
        if 2 <= len(words) <= 5 and 5 <= len(match) <= 50:
            results['names'].append(match)
    
    # Extract class patterns
    class_pattern = r'\b(\d{1,2}[A-Z]{1,4}\d{0,2})\b'
    class_matches = re.findall(class_pattern, line, re.IGNORECASE)
    results['classes'].extend(class_matches)
    
    # Extract date patterns
    date_pattern = r'\b(\d{1,2}[/-]\d{1,2}[/-]\d{4})\b'
    date_matches = re.findall(date_pattern, line)
    results['dates'].extend(date_matches)
    
    # Remove empty lists
    return {k: v for k, v in results.items() if v}

def intelligent_data_extraction(text):
    """Intelligent extraction of student data with advanced recognition"""
    # Normalize text
    normalized_text = unicodedata.normalize('NFKD', text)
    
    extracted_info = {
        'mssv': set(),
        'names': set(),
        'classes': set(),
        'dates': set(),
        'mixed_info': []
    }
    
    # Split into lines and clean
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    
    logger.info(f"Processing {len(lines)} lines for intelligent extraction")
    
    for line_idx, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
            
        logger.debug(f"Processing line {line_idx + 1}: '{line}'")
        
        # Strategy 1: Pure MSSV detection (high confidence)
        if line.isdigit() and 6 <= len(line) <= 10:
            # Additional validation for Vietnamese student ID patterns
            if (len(line) == 8 and line.startswith('20')) or \
               (len(line) == 10 and line.startswith('20')) or \
               (len(line) == 7) or (len(line) == 6):
                extracted_info['mssv'].add(line)
                logger.debug(f"Found pure MSSV: {line}")
                continue
        
        # Strategy 2: Pure Vietnamese name detection (high confidence)
        words = line.split()
        if is_vietnamese_name(line, words):
            extracted_info['names'].add(line)
            logger.debug(f"Found Vietnamese name: {line}")
            continue
        
        # Strategy 3: Class pattern detection
        if is_class_pattern(line):
            extracted_info['classes'].add(line)
            logger.debug(f"Found class: {line}")
            continue
        
        # Strategy 4: Date pattern detection
        if is_date_pattern(line):
            extracted_info['dates'].add(line)
            logger.debug(f"Found date: {line}")
            continue
        
        # Strategy 5: Mixed content parsing
        mixed_results = parse_mixed_content(line)
        if mixed_results:
            extracted_info['mixed_info'].append({
                'line': line,
                'parsed': mixed_results
            })
            # Add parsed items to respective sets
            for item_type, items in mixed_results.items():
                if item_type in extracted_info and isinstance(extracted_info[item_type], set):
                    extracted_info[item_type].update(items)
    
    # Convert sets to lists for JSON serialization
    result = {}
    for key, value in extracted_info.items():
        if isinstance(value, set):
            result[key] = list(value)
        else:
            result[key] = value
    
    logger.info(f"Intelligent extraction complete - MSSV: {len(result['mssv'])}, Names: {len(result['names'])}, Classes: {len(result['classes'])}, Dates: {len(result['dates'])}")
    
    return result

def validate_file(file):
    """Validate uploaded file"""
    if not file or not file.filename:
        return False, "Không có file được chọn"
    
    if file.content_length and file.content_length > 16 * 1024 * 1024:
        return False, f"File '{file.filename}' quá lớn. Kích thước tối đa là 16MB"
    
    allowed_extensions = {'.csv', '.docx', '.xlsx', '.xls', '.txt'}
    file_extension = os.path.splitext(file.filename)[1].lower()
    
    if file_extension not in allowed_extensions:
        return False, f"File '{file.filename}' không được hỗ trợ. Chỉ hỗ trợ: CSV, DOCX, XLSX, TXT"
    
    # Check if required libraries are available
    if file_extension == '.docx' and not DOCX_AVAILABLE:
        return False, "DOCX files không được hỗ trợ. Vui lòng cài đặt python-docx"
    
    if file_extension in ['.xlsx', '.xls'] and not XLSX_AVAILABLE:
        return False, "XLSX files không được hỗ trợ. Vui lòng cài đặt openpyxl"
    
    return True, "OK"

def process_uploaded_files(files):
    """Process multiple uploaded files and extract text content"""
    all_extracted_data = {
        'mssv': [],
        'names': [],
        'classes': [],
        'dates': []
    }
    
    processed_files = []
    
    for file in files:
        if not file or not file.filename:
            continue
            
        # Validate file
        is_valid, error_msg = validate_file(file)
        if not is_valid:
            logger.warning(f"File validation failed: {error_msg}")
            continue
        
        try:
            filename = secure_filename(file.filename)
            file_extension = os.path.splitext(filename)[1].lower()
            
            # Save file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
                file.save(temp_file.name)
                
                # Extract text from file
                file_text = extract_text_from_file(temp_file.name, file_extension)
                if file_text:
                    file_data = intelligent_data_extraction(file_text)
                    
                    # Merge with all data
                    for key in all_extracted_data.keys():
                        all_extracted_data[key].extend(file_data.get(key, []))
                    
                    processed_files.append({
                        'name': filename,
                        'size': os.path.getsize(temp_file.name),
                        'extracted_items': sum(len(items) for items in file_data.values() if isinstance(items, list))
                    })
                    
                    logger.info(f"Processed file: {filename}, extracted {sum(len(items) for items in file_data.values() if isinstance(items, list))} items")
            
            # Clean up temp file
            os.unlink(temp_file.name)
            
        except Exception as e:
            logger.error(f"Error processing file {file.filename}: {e}")
            continue
    
    # Remove duplicates
    for key in all_extracted_data.keys():
        all_extracted_data[key] = list(set(all_extracted_data[key]))
    
    return all_extracted_data, processed_files

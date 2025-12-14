"""
Flask App - HUIT Student Management Dashboard
Version 2.0 - Based on working simple version with full features
"""

try:
    from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
    import pandas as pd
    import json
    import os
    import numpy as np
    import io
    import base64
    from datetime import datetime
    from pathlib import Path
    import logging
    import re
    import unicodedata
except ImportError as e:
    print(f"Import error: {e}")
    print("Please install required packages:")
    print("pip install flask pandas numpy")
    exit(1)

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'huit_student_dashboard_2024'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Global data storage
student_data = None
uploaded_files = []

# Helper functions
def load_sample_data():
    """Load sample data if available"""
    global student_data
    
    # Ưu tiên file dữ liệu HUIT mới từ folder data ở root
    sample_files = [
        '../data/huit_point_student.csv', 
        '../data/sample_data_huit.csv', 
        '../data/sample_data.csv', 
        '../data/data.csv',
        'data/huit_point_student.csv',  # nếu có trong flask_app/data
        'huit_point_student.csv',       # fallback nếu file ở flask_app
        'sample_data.csv'
    ]
    
    for file_path in sample_files:
        if os.path.exists(file_path):
            try:
                student_data = pd.read_csv(file_path)
                logger.info(f"Loaded {len(student_data)} records from {file_path}")
                return True
            except Exception as e:
                logger.warning(f"Error loading {file_path}: {e}")
    
    return False

def prepare_dashboard_data():
    """Prepare data for dashboard"""
    global student_data
    
    if student_data is None or student_data.empty:
        return {
            'total_students': 0,
            'total_classes': 0,
            'average_score': 0,
            'pass_rate': 0,
            'has_data': False,
            'chart_data': {
                'class_distribution': [],
                'score_distribution': [],
                'monthly_trends': []
            }
        }
    
    # Basic stats
    total_students = len(student_data)
    
    # Try different column names for class
    class_col = None
    for col in ['lop', 'class', 'Lớp', 'Class']:
        if col in student_data.columns:
            class_col = col
            break
    
    total_classes = student_data[class_col].nunique() if class_col else 1
    
    # Try different column names for scores
    score_cols = []
    for col in student_data.columns:
        if any(keyword in col.lower() for keyword in ['diem', 'score', 'điểm', 'point']):
            try:
                # Test if column contains numeric data
                test_numeric = pd.to_numeric(student_data[col], errors='coerce')
                if not pd.isna(test_numeric).all():  # If at least some values are numeric
                    score_cols.append(col)
            except:
                continue
    
    average_score = 0
    pass_rate = 0
    
    if score_cols:
        # Use first numeric score column
        main_score_col = score_cols[0]
        logger.info(f"Using score column: {main_score_col}")
        scores = pd.to_numeric(student_data[main_score_col], errors='coerce').dropna()
        
        if len(scores) > 0:
            average_score = float(scores.mean())
            # Count scores >= 5.0 as pass (Vietnamese grading system)
            pass_rate = float((scores >= 5.0).mean() * 100)
            logger.info(f"Calculated average: {average_score:.2f}, pass rate: {pass_rate:.1f}%")
    
    # Chart data for class distribution
    chart_data = {
        'class_distribution': [],
        'score_distribution': [],
        'monthly_trends': []
    }
    
    # Class distribution
    if class_col:
        class_counts = student_data[class_col].value_counts().head(10)
        chart_data['class_distribution'] = [
            {'label': str(cls), 'value': int(count)} 
            for cls, count in class_counts.items()
        ]
    
    # Score distribution
    if score_cols:
        main_score_col = score_cols[0]
        scores = pd.to_numeric(student_data[main_score_col], errors='coerce').dropna()
        if len(scores) > 0:
            bins = [0, 2, 4, 5, 6.5, 8, 10]
            labels = ['Kém (<2)', 'Yếu (2-4)', 'TB- (4-5)', 'TB (5-6.5)', 'Khá (6.5-8)', 'Giỏi (8-10)']
            
            # Create histogram
            hist, _ = np.histogram(scores, bins=bins)
            chart_data['score_distribution'] = [
                {'label': labels[i], 'value': int(count)} 
                for i, count in enumerate(hist)
            ]
    
    return {
        'total_students': total_students,
        'total_classes': total_classes,
        'average_score': round(average_score, 2),
        'pass_rate': round(pass_rate, 1),
        'recent_updates': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'has_data': True,
        'chart_data': chart_data
    }

def advanced_search_students(query, search_type='all'):
    """Advanced search function with improved algorithm"""
    global student_data
    
    if student_data is None or student_data.empty:
        return []
    
    query = str(query).strip().lower()
    if not query:
        return []
    
    matches = []
    
    # Identify important columns
    mssv_cols = []
    name_cols = []
    class_cols = []
    
    for col in student_data.columns:
        col_lower = col.lower()
        if any(keyword in col_lower for keyword in ['mssv', 'id', 'ma_sv', 'student_id', 'mã sinh viên']):
            mssv_cols.append(col)
        elif any(keyword in col_lower for keyword in ['tên', 'name', 'ho_ten', 'fullname', 'họ và tên', 'họ đệm']):
            name_cols.append(col)
        elif any(keyword in col_lower for keyword in ['lớp', 'class', 'lop']):
            class_cols.append(col)
    
    # Search logic based on type
    for idx in student_data.index:
        student = student_data.loc[idx]
        student_dict = student.to_dict()
        match_score = 0
        match_details = []
        
        # Search by MSSV
        if search_type in ['all', 'mssv']:
            for col in mssv_cols:
                value = str(student[col]).strip().lower()
                if value != 'nan' and value:
                    if query in value:
                        if query == value:
                            match_score = max(match_score, 1.0)  # Exact match
                            match_details.append(f"MSSV khớp chính xác: {student[col]}")
                        else:
                            match_score = max(match_score, 0.8)  # Partial match
                            match_details.append(f"MSSV chứa: {student[col]}")
        
        # Search by name
        if search_type in ['all', 'name']:
            # Try to find full name by combining "Họ đệm" and "Tên" columns
            full_name = ""
            ho_dem = ""
            ten = ""
            
            for col in name_cols:
                if 'họ đệm' in col.lower():
                    ho_dem = str(student[col]).strip()
                elif 'tên' in col.lower() and 'họ' not in col.lower():
                    ten = str(student[col]).strip()
            
            if ho_dem and ten:
                full_name = f"{ho_dem} {ten}".strip().lower()
            else:
                # Fallback to individual name columns
                for col in name_cols:
                    value = str(student[col]).strip().lower()
                    if value != 'nan' and value:
                        full_name = value
                        break
            
            if full_name and full_name != 'nan':
                # Exact match
                if query == full_name:
                    match_score = max(match_score, 0.95)
                    match_details.append(f"Tên khớp chính xác: {full_name}")
                # Contains match
                elif query in full_name or full_name in query:
                    match_score = max(match_score, 0.85)
                    match_details.append(f"Tên chứa: {full_name}")
                # Word-by-word match
                else:
                    query_words = query.split()
                    name_words = full_name.split()
                    word_matches = sum(1 for qw in query_words if any(qw in nw for nw in name_words))
                    if word_matches > 0:
                        score = 0.7 * (word_matches / len(query_words))
                        if score > 0.3:
                            match_score = max(match_score, score)
                            match_details.append(f"Tên khớp một phần: {full_name}")
        
        # Search by class
        if search_type in ['all', 'class']:
            for col in class_cols:
                value = str(student[col]).strip().lower()
                if value != 'nan' and value:
                    if query in value:
                        if query == value:
                            match_score = max(match_score, 0.9)
                            match_details.append(f"Lớp khớp chính xác: {student[col]}")
                        else:
                            match_score = max(match_score, 0.7)
                            match_details.append(f"Lớp chứa: {student[col]}")
        
        # Add to results if good match
        if match_score > 0.3:
            student_dict['match_score'] = match_score
            student_dict['match_details'] = match_details
            student_dict['student_index'] = idx
            
            # Clean up display values
            for key, value in student_dict.items():
                if pd.isna(value) or str(value).lower() == 'nan':
                    student_dict[key] = ''
                elif isinstance(value, float) and value.is_integer():
                    student_dict[key] = int(value)
            
            matches.append(student_dict)
    
    # Sort by original data order (student_index) instead of match score
    matches.sort(key=lambda x: x['student_index'])
    
    # Return all matches (no limit)
    return matches

def perform_search(query, filters=None):
    """Perform search on student data"""
    global student_data
    
    if student_data is None or student_data.empty:
        return []
    
    result = student_data.copy()
    
    # Text search
    if query:
        text_cols = []
        for col in student_data.columns:
            if student_data[col].dtype == 'object':
                text_cols.append(col)
        
        if text_cols:
            mask = False
            for col in text_cols:
                mask |= result[col].astype(str).str.contains(query, case=False, na=False)
            result = result[mask]
    
    # Apply filters
    if filters:
        if 'class' in filters and filters['class']:
            class_col = None
            for col in ['lop', 'class', 'Lớp', 'Class']:
                if col in result.columns:
                    class_col = col
                    break
            if class_col:
                result = result[result[class_col] == filters['class']]
        
        if 'min_score' in filters and filters['min_score']:
            score_cols = []
            for col in result.columns:
                if any(keyword in col.lower() for keyword in ['diem', 'score', 'điểm']):
                    if result[col].dtype in ['float64', 'int64']:
                        score_cols.append(col)
            
            if score_cols:
                main_score_col = score_cols[0]
                scores = pd.to_numeric(result[main_score_col], errors='coerce')
                result = result[scores >= float(filters['min_score'])]
    
    return result.head(100).to_dict('records')

# Routes
@app.route('/')
def index():
    """Dashboard page"""
    dashboard_data = prepare_dashboard_data()
    return render_template('dashboard.html', **dashboard_data)

@app.route('/students')
def students():
    """Students list page with search and filter support"""
    global student_data
    
    page = int(request.args.get('page', 1))
    per_page = 20
    
    if student_data is None or student_data.empty:
        return render_template('students.html', 
                             students=[], 
                             total=0, 
                             page=page, 
                             per_page=per_page,
                             has_data=False)
    
    # Get search and filter parameters
    search_query = request.args.get('search', '').strip()
    class_filter = request.args.get('class', '').strip()
    sort_option = request.args.get('sort', 'name')
    
    # Start with all data
    filtered_data = student_data.copy()
    
    # Apply search filter
    if search_query:
        search_query_lower = search_query.lower()
        
        # Search in multiple columns
        search_mask = pd.Series([False] * len(filtered_data))
        
        # Search columns to check
        search_columns = []
        
        # Find relevant columns
        for col in filtered_data.columns:
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in ['tên', 'name', 'ho_ten', 'họ và tên', 'mssv', 'id', 'ma_sv', 'mã sinh viên']):
                search_columns.append(col)
        
        # Search in identified columns
        for col in search_columns:
            try:
                search_mask |= filtered_data[col].astype(str).str.lower().str.contains(search_query_lower, na=False)
            except:
                continue
        
        # If no specific columns found, search in all text columns
        if not search_columns:
            for col in filtered_data.columns:
                if filtered_data[col].dtype == 'object':
                    try:
                        search_mask |= filtered_data[col].astype(str).str.lower().str.contains(search_query_lower, na=False)
                    except:
                        continue
        
        filtered_data = filtered_data[search_mask]
    
    # Apply class filter
    if class_filter:
        class_col = None
        for col in ['Lớp', 'lop', 'class', 'Class']:
            if col in filtered_data.columns:
                class_col = col
                break
        
        if class_col:
            filtered_data = filtered_data[filtered_data[class_col].astype(str).str.strip() == class_filter]
    
    # Apply sorting
    if sort_option == 'name':
        # Sort by name columns
        name_cols = []
        for col in ['Họ và tên', 'Tên', 'Name', 'Ho_ten']:
            if col in filtered_data.columns:
                name_cols.append(col)
                break
        if name_cols:
            filtered_data = filtered_data.sort_values(name_cols[0])
    elif sort_option == 'mssv':
        # Sort by MSSV
        mssv_cols = []
        for col in ['Mã sinh viên', 'MSSV', 'ID', 'Ma_sv']:
            if col in filtered_data.columns:
                mssv_cols.append(col)
                break
        if mssv_cols:
            filtered_data = filtered_data.sort_values(mssv_cols[0])
    elif sort_option == 'class':
        # Sort by class
        class_cols = []
        for col in ['Lớp', 'Class', 'lop']:
            if col in filtered_data.columns:
                class_cols.append(col)
                break
        if class_cols:
            filtered_data = filtered_data.sort_values(class_cols[0])
    elif sort_option == 'score':
        # Sort by score
        score_cols = []
        for col in filtered_data.columns:
            if any(keyword in col.lower() for keyword in ['điểm', 'score', 'diem']):
                score_cols.append(col)
        if score_cols:
            # Use the first score column found
            filtered_data = filtered_data.sort_values(score_cols[0], ascending=False, na_position='last')
    
    # Calculate pagination based on filtered data
    total_filtered = len(filtered_data)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    
    students_list = filtered_data.iloc[start_idx:end_idx].to_dict('records')
    
    # Combine "Họ đệm" and "Tên" into "Họ và tên" for each student
    for student in students_list:
        ho_dem = str(student.get('Họ đệm', '')).strip()
        ten = str(student.get('Tên', '')).strip()
        # Combine them with a space, handle cases where one might be empty
        full_name = f"{ho_dem} {ten}".strip()
        student['Họ và tên'] = full_name
        
        # Remove .0 from student ID if it exists
        if 'Mã sinh viên' in student and student['Mã sinh viên'] is not None:
            student_id = str(student['Mã sinh viên'])
            if student_id.endswith('.0'):
                student['Mã sinh viên'] = student_id[:-2]
    
    total = len(student_data)  # Total in original dataset
    total_pages = (total_filtered + per_page - 1) // per_page  # Pages based on filtered results
    
    # Get column names for display, keeping only the most important columns
    # Define important columns in order of preference
    important_columns = ['Mã sinh viên', 'Giới tính', 'Ngày sinh', 'Lớp']
    
    # Find the best score column (prefer the later Điểm 10 column)
    score_columns = [col for col in student_data.columns if 'điểm 10' in col.lower()]
    if score_columns:
        # Take the last one (usually the cumulative score)
        important_columns.append(score_columns[-1])
    
    # Find the best ranking column
    ranking_columns = [col for col in student_data.columns if 'xếp loại' in col.lower()]
    if ranking_columns:
        important_columns.append(ranking_columns[-1])
    
    # Add notes if available
    if 'Ghi chú' in student_data.columns:
        important_columns.append('Ghi chú')
    
    # Filter to only include columns that exist in the data
    columns = [col for col in important_columns if col in student_data.columns]
    
    # Insert "Họ và tên" at the beginning
    columns.insert(0, 'Họ và tên')
    
    # Create a mapping for column name display
    column_display_names = {}
    for col in columns:
        if 'điểm 10' in col.lower() and col != 'Điểm hệ 10':
            column_display_names[col] = 'Điểm hệ 10'
        elif 'xếp loại' in col.lower() and col != 'Xếp loại':
            column_display_names[col] = 'Xếp loại'
        elif col == 'Mã sinh viên':
            column_display_names[col] = 'MSSV'
        elif col == 'Họ và tên':
            column_display_names[col] = 'Tên'
        else:
            column_display_names[col] = col
    
    # Get unique classes for filter dropdown
    class_col = None
    unique_classes = []
    for col in ['Lớp', 'lop', 'class', 'Class']:
        if col in student_data.columns:
            class_col = col
            break
    
    if class_col:
        unique_classes = sorted(student_data[class_col].dropna().unique().astype(str))
    
    # Define detailed fields for student modal (in order you want them displayed)
    detail_fields = [
        'Mã sinh viên',
        'Họ đệm', 
        'Tên',
        'Giới tính',
        'Ngày sinh',
        'Lớp',
        'Điểm 10',
        'Điểm 4', 
        'Điểm chữ',
        'Xếp loại',
        'Năm thứ',
        'Số TC tích lũy',
        'Ghi chú'
    ]
    
    # Filter to only include fields that exist in the data
    available_detail_fields = [field for field in detail_fields if field in student_data.columns]
    
    # Create display names for detail fields
    detail_field_display_names = {
        'Mã sinh viên': 'Mã sinh viên',
        'Họ đệm': 'Họ đệm',
        'Tên': 'Tên',
        'Giới tính': 'Giới tính',
        'Ngày sinh': 'Ngày sinh',
        'Lớp': 'Lớp',
        'Điểm 10': 'Điểm hệ 10',
        'Điểm 4': 'Điểm hệ 4',
        'Điểm chữ': 'Điểm chữ',
        'Xếp loại': 'Xếp loại',
        'Năm thứ': 'Năm thứ',
        'Số TC tích lũy': 'Số TC tích lũy',
        'Ghi chú': 'Ghi chú'
    }
    
    return render_template('students.html', 
                         students=students_list, 
                         columns=columns,
                         column_display_names=column_display_names,
                         detail_fields=available_detail_fields,
                         detail_field_display_names=detail_field_display_names,
                         total=total,
                         total_students=total,
                         total_filtered=total_filtered,  # Add filtered count
                         page=page,
                         current_page_num=page,
                         total_pages=total_pages,
                         per_page=per_page,
                         has_data=True,
                         # Pass search/filter parameters back to template
                         search_query=search_query,
                         class_filter=class_filter,
                         sort_option=sort_option,
                         unique_classes=unique_classes)

@app.route('/search')
def search():
    """Search page"""
    return render_template('search.html', 
                         has_data=(student_data is not None and not student_data.empty),
                         search_results=[])

@app.route('/data-management')
def data_management():
    """Data management page"""
    global uploaded_files, student_data
    
    # Prepare data info
    data_info = {
        'total_students': 0,
        'total_classes': 0,
        'total_columns': 0,
        'file_size': 'N/A',
        'last_updated': 'Chưa có dữ liệu',
        'file_name': 'Chưa tải file'
    }
    
    if student_data is not None and not student_data.empty:
        data_info['total_students'] = len(student_data)
        data_info['total_columns'] = len(student_data.columns)
        
        # Find class column
        class_col = None
        for col in ['Lớp', 'lop', 'class', 'Class']:
            if col in student_data.columns:
                class_col = col
                break
        
        if class_col:
            data_info['total_classes'] = student_data[class_col].nunique()
        
        # Calculate file size
        try:
            import os
            # Kiểm tra các đường dẫn file có thể có
            possible_files = [
                '../data/huit_point_student.csv', 
                'data/huit_point_student.csv', 
                'huit_point_student.csv'
            ]
            file_size = 0
            current_file = 'N/A'
            
            for file_path in possible_files:
                if os.path.exists(file_path):
                    file_size = os.path.getsize(file_path)
                    current_file = file_path
                    break
            
            if file_size > 1024 * 1024:
                data_info['file_size'] = f"{file_size / (1024 * 1024):.1f} MB"
            elif file_size > 1024:
                data_info['file_size'] = f"{file_size / 1024:.1f} KB"
            else:
                data_info['file_size'] = f"{file_size} bytes"
                
            data_info['file_name'] = current_file
        except:
            data_info['file_size'] = 'N/A'
            data_info['file_name'] = 'Chưa tải file'
        
        data_info['last_updated'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    return render_template('data_management.html', 
                         uploaded_files=uploaded_files,
                         data_info=data_info,
                         has_data=(student_data is not None and not student_data.empty))

@app.route('/statistics')
def statistics():
    """Statistics page"""
    return render_template('statistics.html', has_data=(student_data is not None and not student_data.empty))

@app.route('/advanced-search')
def advanced_search():
    """Advanced search page"""
    return render_template('advanced_search.html', has_data=(student_data is not None and not student_data.empty))

@app.route('/compare')
def compare():
    """Compare page"""
    return render_template('compare.html', has_data=(student_data is not None and not student_data.empty))

@app.route('/test-search')
def test_search():
    """Test search functionality"""
    return render_template('test_search.html')

@app.route('/alerts')
def alerts():
    """Alerts page"""
    return render_template('alerts.html', has_data=(student_data is not None and not student_data.empty))

@app.route('/reports')
def reports():
    """Reports page"""
    return render_template('reports.html', has_data=(student_data is not None and not student_data.empty))

@app.route('/class-comparison')
def class_comparison():
    """Class comparison page"""
    return render_template('class_comparison.html')

# API Routes
@app.route('/api/dashboard-data')
def api_dashboard_data():
    """API endpoint for dashboard data"""
    dashboard_data = prepare_dashboard_data()
    return jsonify(dashboard_data)

@app.route('/api/search', methods=['POST', 'GET'])
def api_search():
    """API endpoint for search - handles both POST and GET"""
    if request.method == 'GET':
        # Handle GET request from search form
        query = request.args.get('q', '').strip()
        search_type = request.args.get('type', 'all')
        
        if not query:
            return jsonify({
                'success': True,
                'results': [],
                'total': 0,
                'query': query
            })
        
        # Advanced search logic
        results = advanced_search_students(query, search_type)
        
        return jsonify({
            'success': True,
            'results': results,
            'total': len(results),
            'query': query,
            'search_type': search_type
        })
    else:
        # Handle POST request (existing functionality)
        data = request.get_json()
        query = data.get('query', '')
        filters = data.get('filters', {})
        
        results = perform_search(query, filters)
        
        return jsonify({
            'results': results,
            'total': len(results)
        })

@app.route('/api/student/<int:student_id>')
def api_student_detail(student_id):
    """API endpoint for student details"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    if student_id >= len(student_data):
        return jsonify({'error': 'Student not found'}), 404
    
    student = student_data.iloc[student_id].to_dict()
    return jsonify(student)

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload"""
    global student_data, uploaded_files
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        try:
            filename = file.filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Read file content
            if filename.lower().endswith('.csv'):
                student_data = pd.read_csv(file)
            else:
                student_data = pd.read_excel(file)
            
            # Store file info
            file_info = {
                'name': filename,
                'upload_time': timestamp,
                'rows': len(student_data),
                'columns': len(student_data.columns)
            }
            uploaded_files.append(file_info)
            
            logger.info(f"Uploaded file: {filename}, Rows: {len(student_data)}")
            
            return jsonify({
                'success': True,
                'message': f'Đã tải lên thành công {len(student_data)} bản ghi',
                'file_info': file_info
            })
            
        except Exception as e:
            logger.error(f"Error processing file: {e}")
            return jsonify({'error': f'Lỗi xử lý file: {str(e)}'}), 500
    
    return jsonify({'error': 'File format not supported'}), 400

@app.route('/api/classes')
def api_classes():
    """API endpoint for class list"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify([])
    
    # Find class column
    class_col = None
    for col in ['lop', 'class', 'Lớp', 'Class']:
        if col in student_data.columns:
            class_col = col
            break
    
    if class_col:
        classes = sorted(student_data[class_col].dropna().unique().tolist())
        return jsonify(classes)
    
    return jsonify([])

@app.route('/api/chart/class_distribution')
def api_chart_class_distribution():
    """API endpoint for class distribution chart"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify({'labels': [], 'values': []})
    
    # Find class column
    class_col = None
    for col in ['lop', 'class', 'Lớp', 'Class']:
        if col in student_data.columns:
            class_col = col
            break
    
    if class_col:
        class_counts = student_data[class_col].value_counts().head(8)  # Top 8 classes
        return jsonify({
            'labels': class_counts.index.tolist(),
            'values': class_counts.values.tolist()
        })
    
    return jsonify({'labels': [], 'values': []})

@app.route('/api/statistics/detailed')
def api_statistics_detailed():
    """API endpoint for detailed statistics"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    # Find score column
    score_cols = []
    for col in student_data.columns:
        if any(keyword in col.lower() for keyword in ['diem', 'score', 'điểm', 'point']):
            try:
                test_numeric = pd.to_numeric(student_data[col], errors='coerce')
                if not pd.isna(test_numeric).all():
                    score_cols.append(col)
            except:
                continue
    
    if not score_cols:
        return jsonify({'error': 'No score data found'}), 404
    
    main_score_col = score_cols[0]
    scores = pd.to_numeric(student_data[main_score_col], errors='coerce').dropna()
    
    # Basic statistics
    stats = {
        'total_students': len(scores),
        'mean': float(scores.mean()),
        'median': float(scores.median()),
        'mode': float(scores.mode().iloc[0]) if not scores.mode().empty else 0,
        'std_dev': float(scores.std()),
        'variance': float(scores.var()),
        'min': float(scores.min()),
        'max': float(scores.max()),
        'range': float(scores.max() - scores.min())
    }
    
    # Quartiles
    quartiles = {
        'q1': float(scores.quantile(0.25)),
        'q2': float(scores.quantile(0.5)),  # median
        'q3': float(scores.quantile(0.75)),
        'iqr': float(scores.quantile(0.75) - scores.quantile(0.25))
    }
    
    # Advanced statistics
    coefficient_variation = (stats['std_dev'] / stats['mean']) * 100 if stats['mean'] != 0 else 0
    
    # Skewness (simple calculation)
    skewness = float(((scores - stats['mean']) ** 3).mean() / (stats['std_dev'] ** 3)) if stats['std_dev'] != 0 else 0
    
    # Grade distribution
    bins = [0, 2, 4, 5, 6.5, 8, 10]
    labels = ['Kém (<2)', 'Yếu (2-4)', 'TB- (4-5)', 'TB (5-6.5)', 'Khá (6.5-8)', 'Giỏi (8-10)']
    hist, _ = np.histogram(scores, bins=bins)
    
    grade_distribution = []
    for i, count in enumerate(hist):
        grade_distribution.append({
            'label': labels[i],
            'count': int(count),
            'percentage': float((count / len(scores)) * 100)
        })
    
    # Pass rates by grade
    pass_rate = float((scores >= 5.0).mean() * 100)
    excellent_rate = float((scores >= 8.0).mean() * 100)
    
    # Outliers detection (IQR method)
    q1, q3 = quartiles['q1'], quartiles['q3']
    iqr = quartiles['iqr']
    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr
    
    outliers = scores[(scores < lower_bound) | (scores > upper_bound)]
    outlier_count = len(outliers)
    
    return jsonify({
        'basic_stats': stats,
        'quartiles': quartiles,
        'advanced_stats': {
            'coefficient_variation': round(coefficient_variation, 2),
            'skewness': round(skewness, 3),
            'pass_rate': round(pass_rate, 1),
            'excellent_rate': round(excellent_rate, 1),
            'outlier_count': outlier_count,
            'outlier_percentage': round((outlier_count / len(scores)) * 100, 1)
        },
        'grade_distribution': grade_distribution
    })

@app.route('/api/statistics/class-analysis')
def api_class_analysis():
    """API endpoint for class-wise analysis"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    # Find class and score columns
    class_col = None
    for col in ['lop', 'class', 'Lớp', 'Class']:
        if col in student_data.columns:
            class_col = col
            break
    
    score_cols = []
    for col in student_data.columns:
        if any(keyword in col.lower() for keyword in ['diem', 'score', 'điểm', 'point']):
            try:
                test_numeric = pd.to_numeric(student_data[col], errors='coerce')
                if not pd.isna(test_numeric).all():
                    score_cols.append(col)
            except:
                continue
    
    if not class_col or not score_cols:
        return jsonify({'error': 'Required columns not found'}), 404
    
    main_score_col = score_cols[0]
    class_analysis = []
    
    for class_name in student_data[class_col].dropna().unique():
        class_data = student_data[student_data[class_col] == class_name]
        class_scores = pd.to_numeric(class_data[main_score_col], errors='coerce').dropna()
        
        if len(class_scores) > 0:
            class_stats = {
                'class_name': str(class_name),
                'student_count': len(class_scores),
                'mean': float(class_scores.mean()),
                'median': float(class_scores.median()),
                'std_dev': float(class_scores.std()),
                'min': float(class_scores.min()),
                'max': float(class_scores.max()),
                'pass_rate': float((class_scores >= 5.0).mean() * 100),
                'excellent_rate': float((class_scores >= 8.0).mean() * 100)
            }
            class_analysis.append(class_stats)
    
    # Sort by mean score
    class_analysis.sort(key=lambda x: x['mean'], reverse=True)
    
    return jsonify(class_analysis)

@app.route('/api/statistics/top-students')
def api_top_students():
    """API endpoint for top performing students"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    # Find required columns
    score_cols = []
    for col in student_data.columns:
        if any(keyword in col.lower() for keyword in ['diem', 'score', 'điểm', 'point']):
            try:
                test_numeric = pd.to_numeric(student_data[col], errors='coerce')
                if not pd.isna(test_numeric).all():
                    score_cols.append(col)
            except:
                continue
    
    if not score_cols:
        return jsonify({'error': 'No score data found'}), 404
    
    main_score_col = score_cols[0]
    
    # Find name column
    name_col = None
    for col in student_data.columns:
        if any(keyword in col.lower() for keyword in ['ten', 'name', 'tên', 'ho_ten', 'họ_tên']):
            name_col = col
            break
    
    # Find class column
    class_col = None
    for col in ['lop', 'class', 'Lớp', 'Class']:
        if col in student_data.columns:
            class_col = col
            break
    
    # Create working dataframe
    work_df = student_data.copy()
    work_df['numeric_score'] = pd.to_numeric(work_df[main_score_col], errors='coerce')
    work_df = work_df.dropna(subset=['numeric_score'])
    
    # Get top 10 students
    top_students = work_df.nlargest(10, 'numeric_score')
    
    top_list = []
    for idx, (_, student) in enumerate(top_students.iterrows()):
        try:
            score_val = float(student['numeric_score'])
        except (ValueError, TypeError):
            score_val = 0.0
            
        try:
            name_val = str(student[name_col]) if name_col else f"Sinh viên {idx + 1}"
        except (KeyError, TypeError):
            name_val = f"Sinh viên {idx + 1}"
            
        try:
            class_val = str(student[class_col]) if class_col else "N/A"
        except (KeyError, TypeError):
            class_val = "N/A"
        
        student_info = {
            'rank': idx + 1,
            'score': score_val,
            'name': name_val,
            'class': class_val
        }
        top_list.append(student_info)
    
    # Get low performers (bottom 10 with scores < 5.0)
    low_performers_df = work_df[work_df['numeric_score'] < 5.0]
    if not low_performers_df.empty:
        low_performers = low_performers_df.iloc[low_performers_df['numeric_score'].argsort()[:10]]
    else:
        low_performers = low_performers_df
    
    low_list = []
    for idx, (_, student) in enumerate(low_performers.iterrows()):
        student_info = {
            'rank': idx + 1,
            'score': float(student['numeric_score']),
            'name': str(student[name_col]) if name_col else f"Sinh viên {idx + 1}",
            'class': str(student[class_col]) if class_col else "N/A"
        }
        low_list.append(student_info)
    
    return jsonify({
        'top_students': top_list,
        'low_performers': low_list
    })

@app.route('/api/statistics/recommendations')
def api_statistics_recommendations():
    """API endpoint for AI-based recommendations"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify({'recommendations': []})
    
    recommendations = []
    
    # Find score column
    score_cols = []
    for col in student_data.columns:
        if any(keyword in col.lower() for keyword in ['diem', 'score', 'điểm', 'point']):
            try:
                test_numeric = pd.to_numeric(student_data[col], errors='coerce')
                if not pd.isna(test_numeric).all():
                    score_cols.append(col)
            except:
                continue
    
    if score_cols:
        main_score_col = score_cols[0]
        scores = pd.to_numeric(student_data[main_score_col], errors='coerce').dropna()
        
        mean_score = scores.mean()
        std_dev = scores.std()
        pass_rate = (scores >= 5.0).mean() * 100
        
        # Generate recommendations based on data analysis
        if pass_rate < 70:
            recommendations.append({
                'type': 'warning',
                'title': 'Tỷ lệ qua môn thấp',
                'message': f'Chỉ có {pass_rate:.1f}% sinh viên đạt điểm qua môn. Cần có biện pháp hỗ trợ học tập.',
                'priority': 'high'
            })
        
        if std_dev > 2.0:
            recommendations.append({
                'type': 'info',
                'title': 'Sự phân tán điểm cao',
                'message': f'Độ lệch chuẩn {std_dev:.2f} cho thấy sự chênh lệch lớn giữa sinh viên. Cần phân loại và hỗ trợ theo nhóm.',
                'priority': 'medium'
            })
        
        if mean_score < 5.0:
            recommendations.append({
                'type': 'error',
                'title': 'Điểm trung bình thấp',
                'message': f'Điểm trung bình {mean_score:.2f} dưới mức qua môn. Cần xem xét lại phương pháp giảng dạy.',
                'priority': 'high'
            })
        elif mean_score > 8.0:
            recommendations.append({
                'type': 'success',
                'title': 'Chất lượng học tập tốt',
                'message': f'Điểm trung bình {mean_score:.2f} ở mức cao. Cần duy trì và phát huy.',
                'priority': 'low'
            })
        
        # Class-based recommendations
        class_col = None
        for col in ['lop', 'class', 'Lớp', 'Class']:
            if col in student_data.columns:
                class_col = col
                break
        
        if class_col:
            class_means = student_data.groupby(class_col)[main_score_col].apply(
                lambda x: pd.to_numeric(x, errors='coerce').mean()
            ).dropna()
            
            if len(class_means) > 1:
                worst_class = class_means.idxmin()
                best_class = class_means.idxmax()
                
                if class_means[worst_class] < mean_score - std_dev:
                    recommendations.append({
                        'type': 'warning',
                        'title': f'Lớp {worst_class} cần hỗ trợ',
                        'message': f'Lớp {worst_class} có điểm trung bình {class_means[worst_class]:.2f}, thấp hơn đáng kể so với mức chung.',
                        'priority': 'medium'
                    })
                
                if class_means[best_class] > mean_score + std_dev:
                    recommendations.append({
                        'type': 'success',
                        'title': f'Lớp {best_class} xuất sắc',
                        'message': f'Lớp {best_class} đạt điểm trung bình {class_means[best_class]:.2f}, có thể làm mô hình cho các lớp khác.',
                        'priority': 'low'
                    })
    
    return jsonify({'recommendations': recommendations})

@app.route('/api/chart/score_histogram')
def api_chart_score_histogram():
    """API endpoint for score histogram chart"""
    global student_data
    
    if student_data is None or student_data.empty:
        return jsonify({'bins': [], 'counts': []})
    
    # Find score columns
    score_cols = []
    for col in student_data.columns:
        if any(keyword in col.lower() for keyword in ['diem', 'score', 'điểm', 'point']):
            try:
                # Test if column contains numeric data
                test_numeric = pd.to_numeric(student_data[col], errors='coerce')
                if not pd.isna(test_numeric).all():  # If at least some values are numeric
                    score_cols.append(col)
            except:
                continue
    
    if score_cols:
        main_score_col = score_cols[0]
        scores = pd.to_numeric(student_data[main_score_col], errors='coerce').dropna()
        
        if len(scores) > 0:
            # Create histogram bins
            bins = [0, 2, 4, 5, 6.5, 8, 10]
            labels = ['<2 (Kém)', '2-4 (Yếu)', '4-5 (TB-)', '5-6.5 (TB)', '6.5-8 (Khá)', '8-10 (Giỏi)']
            
            # Calculate histogram
            hist, bin_edges = np.histogram(scores, bins=bins)
            
            return jsonify({
                'labels': labels,
                'counts': hist.tolist(),
                'bins': bin_edges.tolist()
            })
    
    return jsonify({'labels': [], 'counts': [], 'bins': []})

# ================================
# ADVANCED SEARCH FEATURES
# ================================

# Import our custom modules
from file_processor import intelligent_data_extraction, process_uploaded_files, validate_file, secure_filename
from student_search import advanced_student_search, identify_database_columns

@app.route('/api/advanced-search', methods=['POST'])
def api_advanced_search():
    """API endpoint for advanced search with text input - EXACT MATCH ONLY"""
    try:
        data = request.get_json()
        query = data.get('query', '').strip()
        
        if not query:
            return jsonify({
                'success': True,
                'matches': [],
                'total_input_items': 0,
                'extracted_data': {}
            })
        
        # Extract data intelligently - split by lines and clean
        input_lines = [line.strip() for line in query.split('\n') if line.strip()]
        
        # Perform advanced search with exact match only
        matches = advanced_student_search(input_lines, student_data)
        
        return jsonify({
            'success': True,
            'matches': matches,
            'total_input_items': len(input_lines),
            'extracted_data': {'inputs': input_lines}
        })
        
    except Exception as e:
        logger.error(f"Error in advanced search: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/advanced-search/upload', methods=['POST'])
def api_advanced_search_upload():
    """API endpoint for advanced search with file upload - EXACT MATCH ONLY"""
    try:
        files = request.files.getlist('files')
        text_input = request.form.get('text_input', '')
        
        # Combine all input data
        all_inputs = []
        
        # Process text input
        if text_input.strip():
            input_lines = [line.strip() for line in text_input.split('\n') if line.strip()]
            all_inputs.extend(input_lines)
        
        # Process uploaded files (FIXED - avoid duplicates)
        if files:
            for file in files:
                if file and file.filename:
                    try:
                        if file.filename.endswith('.csv'):
                            content = file.read().decode('utf-8')
                            lines = [line.strip() for line in content.split('\n') if line.strip()]
                            
                            # Xử lý CSV thông minh - kiểm tra xem có phải file có header không
                            csv_data = []
                            header_found = False
                            
                            for line in lines:
                                if not line:
                                    continue
                                    
                                # Skip header line
                                if any(keyword in line.lower() for keyword in ['họ tên', 'mssv', 'name', 'student']):
                                    header_found = True
                                    continue
                                
                                # Nếu có dấu phẩy trong dòng và có header, chỉ lấy dữ liệu cần thiết
                                if ',' in line and header_found:
                                    parts = [part.strip() for part in line.split(',') if part.strip()]
                                    # Chỉ lấy phần đầu tiên (tên) hoặc phần thứ hai (MSSV) làm input
                                    # Tránh lấy cả hai để không tạo duplicate
                                    if len(parts) >= 2:
                                        # Ưu tiên MSSV vì chính xác hơn
                                        mssv_part = parts[1].strip()
                                        if mssv_part.isdigit() and len(mssv_part) >= 6:
                                            csv_data.append(mssv_part)
                                        else:
                                            # Nếu không phải MSSV, lấy tên
                                            csv_data.append(parts[0].strip())
                                    else:
                                        csv_data.append(parts[0].strip())
                                else:
                                    # Dòng không có dấu phẩy hoặc không có header
                                    csv_data.append(line)
                            
                            all_inputs.extend(csv_data)
                        elif file.filename.endswith('.txt'):
                            content = file.read().decode('utf-8')
                            lines = [line.strip() for line in content.split('\n') if line.strip()]
                            all_inputs.extend(lines)
                    except Exception as e:
                        logger.warning(f"Error processing file {file.filename}: {e}")
        
        # Remove duplicates
        all_inputs = list(set(all_inputs))
        
        # Perform advanced search with exact match only
        matches = advanced_student_search(all_inputs, student_data)
        
        return jsonify({
            'success': True,
            'matches': matches,
            'total_input_items': len(all_inputs),
            'extracted_data': {'inputs': all_inputs}
        })
        
    except Exception as e:
        logger.error(f"Error in advanced search upload: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ================================
# DATA COMPARISON FEATURES (Updated)
# ================================

import re
try:
    import docx
except ImportError:
    docx = None
    logger.warning("python-docx not installed. DOCX file support disabled.")

try:
    import openpyxl
except ImportError:
    openpyxl = None
    logger.warning("openpyxl not installed. XLSX file support disabled.")

try:
    from werkzeug.utils import secure_filename
except ImportError:
    def secure_filename(filename):
        """Fallback secure_filename implementation"""
        import re
        filename = filename.replace(' ', '_')
        return re.sub(r'[^a-zA-Z0-9._-]', '', filename)
    
import tempfile
from difflib import SequenceMatcher

def extract_text_from_file(file_path, file_extension):
    """Extract text from various file formats"""
    try:
        if file_extension in ['.txt', '.csv']:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif file_extension == '.docx':
            if docx is None:
                logger.warning("DOCX support not available")
                return ""
            doc = docx.Document(file_path)
            return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
        elif file_extension in ['.xlsx', '.xls']:
            if openpyxl is None:
                logger.warning("XLSX support not available")
                return ""
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else '' for cell in row]
                    text_content.append('\t'.join(row_text))
            return '\n'.join(text_content)
        else:
            return ""
    except Exception as e:
        logger.error(f"Error extracting text from {file_path}: {e}")
        return ""

def extract_student_info(text):
    """Extract student names and IDs from text - OPTIMIZED for precision"""
    import unicodedata
    
    # Normalize text - remove special chars, extra spaces  
    normalized_text = unicodedata.normalize('NFKD', text)
    
    extracted_info = {'mssv': set(), 'name': set()}
    
    # Split text into lines and process each line individually
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    
    logger.info(f"Processing {len(lines)} lines of input text")
    
    for line_idx, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
            
        logger.debug(f"Processing line {line_idx + 1}: '{line}'")
        
        # Strategy 1: Check if entire line is a student ID
        if line.isdigit() and 6 <= len(line) <= 10:
            # Additional validation for Vietnamese student ID patterns
            if (len(line) == 8 and line.startswith('20')) or \
               (len(line) == 10 and line.startswith('20')) or \
               (6 <= len(line) <= 10):
                extracted_info['mssv'].add(line)
                logger.debug(f"Found MSSV: {line}")
                continue
        
        # Strategy 2: Check if entire line is a Vietnamese name
        words = line.split()
        if (len(words) >= 2 and  # At least 2 words (first name + last name)
            len(words) <= 5 and  # Not more than 5 words 
            not any(char.isdigit() for char in line) and  # No digits
            5 <= len(line) <= 40 and  # Reasonable length
            all(len(word) >= 1 for word in words) and  # Each word at least 1 char (allowing single letter names)
            not any(word.lower() in ['file', 'data', 'csv', 'excel', 'document', 'text'] for word in words)):  # No tech words
            
            # Check for Vietnamese name characteristics
            has_vietnamese_chars = any(char in 'àáâãèéêìíòóôõùúăđĩũơưăạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵýỷỹ' for char in line.lower())
            has_proper_caps = all(word[0].isupper() and (len(word) == 1 or word[1:].islower()) for word in words)
            is_all_caps = all(word.isupper() for word in words)
            
            # Also check for common Vietnamese names patterns
            has_common_vn_chars = any(char in 'ă ô ư đ' for char in line.lower())
            
            if has_vietnamese_chars or has_proper_caps or is_all_caps or has_common_vn_chars:
                extracted_info['name'].add(line)
                logger.debug(f"Found name: {line}")
                continue
        
        # Strategy 3: Look for patterns within the line (for mixed content)
        # MSSV patterns (more restrictive)
        mssv_patterns = [
            r'\b(20\d{8})\b',      # 10-digit starting with 20
            r'\b(20\d{6})\b',      # 8-digit starting with 20  
            r'\b(\d{6,10})\b'      # General 6-10 digit pattern
        ]
        
        for pattern in mssv_patterns:
            matches = re.findall(pattern, line)
            for match in matches:
                if 6 <= len(match) <= 10:
                    extracted_info['mssv'].add(match)
                    logger.debug(f"Found MSSV in line: {match}")
        
        # Name patterns within line (more restrictive)
        # Only look for names if the line contains mixed content
        if any(char.isdigit() for char in line):  # Mixed content
            name_pattern = r'\b([A-ZÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠƯĂẠẢẤẦẨẪẬẮẰẲẴẶẸẺẼẾỀỂỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪỬỮỰỲỴÝỶỸ][a-zàáâãèéêìíòóôõùúăđĩũơưăạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵýỷỹ]+(?:\s+[A-ZÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠƯĂẠẢẤẦẨẪẬẮẰẲẴẶẸẺẼẾỀỂỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪỬỮỰỲỴÝỶỸ][a-zàáâãèéêìíòóôõùúăđĩũơưăạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵýỷỹ]+)+)\b'
            name_matches = re.findall(name_pattern, line)
            for match in name_matches:
                words = match.split()
                if (2 <= len(words) <= 5 and
                    5 <= len(match) <= 40 and
                    all(len(word) >= 2 for word in words)):
                    extracted_info['name'].add(match)
                    logger.debug(f"Found name in mixed line: {match}")
    
    logger.info(f"Extraction complete - MSSV: {len(extracted_info['mssv'])}, Names: {len(extracted_info['name'])}")
    logger.debug(f"MSVs found: {list(extracted_info['mssv'])}")
    logger.debug(f"Names found: {list(extracted_info['name'])}")
    
    return extracted_info

def fuzzy_match_student(input_text, student_record, threshold=0.6):
    """Use fuzzy matching to find student information"""
    global student_data
    
    input_text = str(input_text).lower().strip()
    
    # Check exact MSSV match first
    for col in student_record.index:
        if any(keyword in col.lower() for keyword in ['mssv', 'id', 'ma_sv', 'student_id']):
            if str(student_record[col]).lower().strip() == input_text:
                return 1.0, 'mssv_exact'
    
    # Check fuzzy name match
    for col in student_record.index:
        if any(keyword in col.lower() for keyword in ['ten', 'name', 'ho_ten', 'fullname']):
            student_name = str(student_record[col]).lower().strip()
            similarity = SequenceMatcher(None, input_text, student_name).ratio()
            if similarity >= threshold:
                return similarity, 'name_fuzzy'
    
    return 0.0, 'no_match'

def find_matching_students(extracted_info):
    """Find matching students in database based on extracted information - OPTIMIZED VERSION"""
    global student_data
    
    if student_data is None or student_data.empty:
        return []
    
    # Calculate total input count for limiting results
    total_input_count = len(extracted_info['mssv']) + len(extracted_info['name'])
    logger.info(f"Input items: {total_input_count} (MSSV: {len(extracted_info['mssv'])}, Names: {len(extracted_info['name'])})")
    
    # Identify column names
    mssv_cols = []
    ho_dem_cols = []
    ten_cols = []
    
    for col in student_data.columns:
        col_lower = col.lower().replace(' ', '_')
        # MSSV columns
        if any(keyword in col_lower for keyword in ['mssv', 'id', 'ma_sv', 'student_id', 'ma_sinh_vien', 'mã_sinh_viên']):
            mssv_cols.append(col)
        # Separate first and last name columns
        elif 'họ_đệm' in col_lower or 'ho_dem' in col_lower:
            ho_dem_cols.append(col)
        elif col_lower == 'tên' or col_lower == 'ten':
            ten_cols.append(col)
    
    logger.info(f"Database columns - MSSV: {mssv_cols}, Ho_dem: {ho_dem_cols}, Ten: {ten_cols}")
    
    # Use a dictionary to store unique matches by student ID to prevent duplicates
    unique_students = {}
    processed_inputs = set()  # Track which inputs we've found matches for
    
    # Phase 1: Search by MSSV (EXACT match only for precision)
    for mssv in extracted_info['mssv']:
        if mssv in processed_inputs:
            continue
            
        mssv_clean = str(mssv).strip()
        match_found = False
        
        for idx, student in student_data.iterrows():
            if match_found:
                break
                
            for mssv_col in mssv_cols:
                student_mssv = str(student[mssv_col]).replace('.0', '').strip()
                
                # ONLY exact match for MSSV to ensure precision
                if student_mssv == mssv_clean:
                    student_info = student.to_dict()
                    student_info['match_score'] = 1.0
                    student_info['match_type'] = 'mssv_exact'
                    student_info['input_value'] = mssv
                    student_info['match_field'] = mssv_col
                    
                    unique_students[student_mssv] = student_info
                    processed_inputs.add(mssv)
                    match_found = True
                    break
    
    # Phase 2: Search by name (EXACT and HIGH fuzzy match only)
    for name in extracted_info['name']:
        if name in processed_inputs:
            continue
            
        name_clean = str(name).strip().lower()
        match_found = False
        best_match = None
        best_score = 0
        
        for idx, student in student_data.iterrows():
            if match_found:
                break
                
            # Get student ID for deduplication
            student_id = None
            for mssv_col in mssv_cols:
                student_id = str(student[mssv_col]).replace('.0', '').strip()
                break
            
            # Skip if we already found this student via MSSV
            if student_id and student_id in unique_students:
                continue
            
            # Try combining ho_dem + ten for precise matching
            if ho_dem_cols and ten_cols:
                for ho_dem_col in ho_dem_cols:
                    for ten_col in ten_cols:
                        ho_dem = str(student[ho_dem_col]).strip() if pd.notna(student[ho_dem_col]) else ""
                        ten = str(student[ten_col]).strip() if pd.notna(student[ten_col]) else ""
                        full_name = f"{ho_dem} {ten}".strip().lower()
                        
                        # Exact match with combined name
                        if full_name == name_clean:
                            student_info = student.to_dict()
                            student_info['match_score'] = 1.0
                            student_info['match_type'] = 'name_exact_combined'
                            student_info['input_value'] = name
                            student_info['match_field'] = f"{ho_dem_col} + {ten_col}"
                            
                            if student_id:
                                unique_students[student_id] = student_info
                                processed_inputs.add(name)
                                match_found = True
                                break
                        
                        # High threshold fuzzy match (85% similarity)
                        similarity = SequenceMatcher(None, name_clean, full_name).ratio()
                        if similarity >= 0.85 and similarity > best_score:
                            best_match = {
                                'student_info': student.to_dict(),
                                'match_score': similarity,
                                'match_type': 'name_fuzzy_combined',
                                'input_value': name,
                                'match_field': f"{ho_dem_col} + {ten_col}",
                                'student_id': student_id
                            }
                            best_score = similarity
                    
                    if match_found:
                        break
        
        # If no exact match found but we have a good fuzzy match
        if not match_found and best_match and best_score >= 0.85:
            student_info = best_match['student_info']
            student_info['match_score'] = best_match['match_score']
            student_info['match_type'] = best_match['match_type']
            student_info['input_value'] = best_match['input_value']
            student_info['match_field'] = best_match['match_field']
            
            if best_match['student_id']:
                unique_students[best_match['student_id']] = student_info
                processed_inputs.add(name)
    
    # Convert to list and clean up data
    final_matches = []
    for student_id, match in unique_students.items():
        # Clean up the data for display
        cleaned_match = {}
        for key, value in match.items():
            if key not in ['match_score', 'match_type', 'input_value', 'match_field']:
                # Clean numeric values
                if str(value).endswith('.0'):
                    cleaned_match[key] = str(value).replace('.0', '')
                else:
                    cleaned_match[key] = value
            else:
                cleaned_match[key] = value
        
        final_matches.append(cleaned_match)
    
    # Sort by match score (highest first)
    final_matches.sort(key=lambda x: x['match_score'], reverse=True)
    
    # Limit results to input count - return exactly as many matches as input items
    limited_matches = final_matches[:total_input_count] if total_input_count > 0 else final_matches
    
    logger.info(f"Found {len(final_matches)} total matches, returning {len(limited_matches)} (limited to input count)")
    logger.info(f"Processed inputs: {len(processed_inputs)}/{total_input_count}")
    
    return limited_matches

@app.route('/api/compare/upload', methods=['POST'])
def api_compare_upload():
    """Handle file upload for comparison"""
    try:
        files = request.files.getlist('files')
        text_input = request.form.get('text_input', '')
        
        all_extracted_info = {'mssv': set(), 'name': set()}
        
        # Process text input
        if text_input.strip():
            text_info = extract_student_info(text_input)
            all_extracted_info['mssv'].update(text_info['mssv'])
            all_extracted_info['name'].update(text_info['name'])
        
        # Process uploaded files
        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                file_extension = os.path.splitext(filename)[1].lower()
                
                # Save file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
                    file.save(temp_file.name)
                    
                    # Extract text from file
                    file_text = extract_text_from_file(temp_file.name, file_extension)
                    file_info = extract_student_info(file_text)
                    
                    all_extracted_info['mssv'].update(file_info['mssv'])
                    all_extracted_info['name'].update(file_info['name'])
                
                # Clean up temp file
                os.unlink(temp_file.name)
        
        # Convert sets to lists for JSON serialization
        extracted_data = {
            'mssv': list(all_extracted_info['mssv']),
            'name': list(all_extracted_info['name'])
        }
        
        # Calculate total input items
        total_input_items = len(extracted_data['mssv']) + len(extracted_data['name'])
        
        # Find matching students
        matches = find_matching_students(all_extracted_info)
        
        return jsonify({
            'success': True,
            'extracted_info': extracted_data,
            'matches': matches,
            'total_matches': len(matches),
            'total_input_items': total_input_items
        })
        
    except Exception as e:
        logger.error(f"Error in compare upload: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/compare/search', methods=['POST'])
def api_compare_search():
    """Search for specific student information"""
    try:
        data = request.get_json()
        query = data.get('query', '').strip()
        
        if not query:
            return jsonify({'matches': [], 'success': True})
        
        # Simple but effective search
        matches = []
        search_lines = [line.strip() for line in query.split('\n') if line.strip()]
        
        for line in search_lines:
            line_matches = search_student_simple(line)
            matches.extend(line_matches)
        
        # Remove duplicates
        unique_matches = []
        seen = set()
        
        for match in matches:
            # Create unique key from MSSV or name
            key = None
            for col_name, value in match.items():
                if any(id_word in col_name.lower() for id_word in ['mssv', 'id', 'ma_sv']):
                    key = f"mssv_{str(value).strip()}"
                    break
            
            if not key:
                for col_name, value in match.items():
                    if any(name_word in col_name.lower() for name_word in ['ten', 'name', 'ho_ten']):
                        key = f"name_{str(value).strip()}"
                        break
            
            if key and key not in seen:
                seen.add(key)
                match['input_value'] = line
                unique_matches.append(match)
        
        return jsonify({
            'success': True,
            'matches': unique_matches,
            'query': query
        })
        
    except Exception as e:
        logger.error(f"Error in compare search: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def search_student_simple(input_text):
    """Simple but effective student search"""
    global student_data
    
    if student_data is None or student_data.empty:
        return []
    
    input_text = input_text.strip()
    if not input_text:
        return []
    
    matches = []
    input_lower = input_text.lower()
    
    # Check if input looks like MSSV (all digits, 6-10 characters)
    is_mssv = input_text.isdigit() and 6 <= len(input_text) <= 10
    
    for idx in student_data.index:
        row_dict = student_data.loc[idx].to_dict()
        match_score = 0
        match_field = None
        
        # Search in all columns
        for col_name, value in row_dict.items():
            if pd.isna(value):
                continue
                
            value_str = str(value).strip().lower()
            
            # MSSV exact match gets highest priority
            if is_mssv and any(id_word in col_name.lower() for id_word in ['mssv', 'id', 'ma_sv']):
                if value_str == input_text:
                    match_score = 1.0
                    match_field = col_name
                    break
                elif input_text in value_str:
                    match_score = 0.9
                    match_field = col_name
            
            # Name matching
            elif any(name_word in col_name.lower() for name_word in ['ten', 'name', 'ho_ten', 'fullname']):
                if input_lower == value_str:
                    match_score = max(match_score, 0.95)
                    match_field = col_name
                elif input_lower in value_str or value_str in input_lower:
                    match_score = max(match_score, 0.8)
                    match_field = col_name
                elif all(word in value_str for word in input_lower.split()):
                    match_score = max(match_score, 0.7)
                    match_field = col_name
            
            # General field matching
            else:
                if input_lower == value_str:
                    match_score = max(match_score, 0.6)
                    match_field = col_name
                elif input_lower in value_str:
                    match_score = max(match_score, 0.5)
                    match_field = col_name
        
        # Add to matches if score is good enough
        if match_score >= 0.5:
            row_dict['match_score'] = match_score
            row_dict['match_field'] = match_field
            matches.append(row_dict)
    
    # Sort by match score (return all matches)
    matches.sort(key=lambda x: x['match_score'], reverse=True)
    return matches  # Return all matches

# ================================
# STUDENT MANAGEMENT FEATURES
# ================================
    matches.sort(key=lambda x: x['match_score'], reverse=True)
    return matches[:10]  # Top 10 matches

# ================================
# STUDENT MANAGEMENT FEATURES
# ================================

@app.route('/api/students/add', methods=['POST'])
def api_add_student():
    """Add a new student"""
    global student_data
    
    try:
        # Get form data
        mssv = request.form.get('mssv', '').strip()
        name = request.form.get('name', '').strip()
        class_name = request.form.get('class', '').strip()
        score = request.form.get('score', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        notes = request.form.get('notes', '').strip()
        
        if not mssv or not name:
            return jsonify({
                'success': False,
                'error': 'MSSV và tên sinh viên là bắt buộc'
            }), 400
        
        # Check for duplicates
        if student_data is not None and not student_data.empty:
            # Find MSSV column
            mssv_col = None
            for col in student_data.columns:
                if any(keyword in col.lower() for keyword in ['mssv', 'id', 'ma_sv', 'student_id']):
                    mssv_col = col
                    break
            
            if mssv_col and mssv in student_data[mssv_col].astype(str).values:
                return jsonify({
                    'success': False,
                    'error': f'MSSV {mssv} đã tồn tại'
                }), 400
        
        # Create new student record
        new_student = {
            'MSSV': mssv,
            'Họ và tên': name,
            'Lớp': class_name,
            'Điểm 10': float(score) if score else None,
            'Email': email,
            'Số điện thoại': phone,
            'Ghi chú': notes
        }
        
        # Add to student_data
        if student_data is None or student_data.empty:
            student_data = pd.DataFrame([new_student])
        else:
            # Ensure all columns exist
            for col in new_student.keys():
                if col not in student_data.columns:
                    student_data[col] = None
            
            # Add new row
            new_row_df = pd.DataFrame([new_student])
            student_data = pd.concat([student_data, new_row_df], ignore_index=True)
        
        logger.info(f"Added new student: {name} ({mssv})")
        
        return jsonify({
            'success': True,
            'message': f'Đã thêm sinh viên {name} thành công'
        })
        
    except Exception as e:
        logger.error(f"Error adding student: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/students/import', methods=['POST'])
def api_import_students():
    """Process file import and return preview data"""
    try:
        files = request.files.getlist('files')
        has_header = request.form.get('hasHeader', 'true').lower() == 'true'
        
        if not files:
            return jsonify({
                'success': False,
                'error': 'Không có file được tải lên'
            }), 400
        
        all_data = []
        
        for file in files:
            if file and file.filename:
                filename = file.filename.lower()
                
                try:
                    if filename.endswith('.csv'):
                        df = pd.read_csv(file, header=0 if has_header else None)
                    elif filename.endswith(('.xlsx', '.xls')):
                        df = pd.read_excel(file, header=0 if has_header else None)
                    else:
                        continue
                    
                    # If no header, assign generic column names
                    if not has_header:
                        df.columns = [f'Column_{i+1}' for i in range(len(df.columns))]
                    
                    # Convert to records
                    records = df.fillna('').to_dict('records')
                    all_data.extend(records)
                    
                except Exception as e:
                    logger.error(f"Error processing file {file.filename}: {e}")
                    return jsonify({
                        'success': False,
                        'error': f'Lỗi xử lý file {file.filename}: {str(e)}'
                    }), 400
        
        if not all_data:
            return jsonify({
                'success': False,
                'error': 'Không tìm thấy dữ liệu hợp lệ trong file'
            }), 400
        
        return jsonify({
            'success': True,
            'data': all_data,
            'count': len(all_data)
        })
        
    except Exception as e:
        logger.error(f"Error in import students: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/students/import/confirm', methods=['POST'])
def api_confirm_import():
    """Confirm and execute student import"""
    global student_data
    
    try:
        data = request.get_json()
        import_data = data.get('data', [])
        skip_duplicates = data.get('skipDuplicates', True)
        
        if not import_data:
            return jsonify({
                'success': False,
                'error': 'Không có dữ liệu để import'
            }), 400
        
        imported_count = 0
        skipped_count = 0
        
        for record in import_data:
            # Try to extract student info from record
            mssv = None
            name = None
            ho_dem = None
            ten = None
            
            # Find MSSV and name from various possible column names
            for key, value in record.items():
                key_lower = key.lower()
                if any(keyword in key_lower for keyword in ['mssv', 'id', 'ma_sv', 'student_id']):
                    mssv = str(value).strip()
                elif any(keyword in key_lower for keyword in ['ten', 'name', 'ho_ten', 'họ_tên', 'fullname']):
                    name = str(value).strip()
                elif 'họ đệm' in key_lower or 'ho_dem' in key_lower:
                    ho_dem = str(value).strip()
                elif key_lower == 'tên' or key_lower == 'ten':
                    ten = str(value).strip()
            
            # Combine "Họ đệm" and "Tên" if they exist separately
            if ho_dem and ten and not name:
                name = f"{ho_dem} {ten}".strip()
            elif not name and (ho_dem or ten):
                name = (ho_dem if ho_dem else ten)
            
            # Skip if missing required fields
            if not mssv or not name:
                skipped_count += 1
                continue
            
            # Check for duplicates if skip_duplicates is enabled
            if skip_duplicates and student_data is not None and not student_data.empty:
                mssv_col = None
                for col in student_data.columns:
                    if any(keyword in col.lower() for keyword in ['mssv', 'id', 'ma_sv', 'student_id']):
                        mssv_col = col
                        break
                
                if mssv_col and mssv in student_data[mssv_col].astype(str).values:
                    skipped_count += 1
                    continue
            
            # Prepare student record with standardized column names
            new_student = {}
            
            # Set the combined name first
            new_student['Họ và tên'] = name
            
            for key, value in record.items():
                # Map to standardized column names
                key_lower = key.lower()
                if any(keyword in key_lower for keyword in ['mssv', 'id', 'ma_sv', 'student_id']):
                    new_student['MSSV'] = str(value).strip()
                elif any(keyword in key_lower for keyword in ['lop', 'class', 'lớp']):
                    new_student['Lớp'] = str(value).strip()
                elif any(keyword in key_lower for keyword in ['diem', 'score', 'điểm', 'point']):
                    try:
                        new_student['Điểm 10'] = float(value) if value else None
                    except:
                        new_student['Điểm 10'] = None
                elif any(keyword in key_lower for keyword in ['email', 'mail']):
                    new_student['Email'] = str(value).strip()
                elif any(keyword in key_lower for keyword in ['phone', 'sdt', 'số_điện_thoại']):
                    new_student['Số điện thoại'] = str(value).strip()
                elif key_lower not in ['họ đệm', 'ho_dem', 'tên', 'ten', 'ten', 'name', 'ho_ten', 'họ_tên', 'fullname']:
                    # Keep original column name for other fields, but skip the name fields we already combined
                    new_student[key] = str(value).strip()
            
            # Add to student_data
            if student_data is None or student_data.empty:
                student_data = pd.DataFrame([new_student])
            else:
                # Ensure all columns exist
                for col in new_student.keys():
                    if col not in student_data.columns:
                        student_data[col] = None
                
                # Add new row
                new_row_df = pd.DataFrame([new_student])
                student_data = pd.concat([student_data, new_row_df], ignore_index=True)
            
            imported_count += 1
        
        logger.info(f"Import completed: {imported_count} imported, {skipped_count} skipped")
        
        return jsonify({
            'success': True,
            'imported': imported_count,
            'skipped': skipped_count,
            'message': f'Import thành công {imported_count} sinh viên' + 
                      (f', bỏ qua {skipped_count} bản ghi' if skipped_count > 0 else '')
        })
        
    except Exception as e:
        logger.error(f"Error confirming import: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/students/template')
def api_download_template():
    """Download student data template"""
    try:
        template_type = request.args.get('type', 'csv')
        
        # Create template data with simplified structure
        template_data = {
            'Mã sinh viên': ['21520001', '21520002', '21520003'],
            'Họ và tên': ['Nguyễn Văn A', 'Trần Thị B', 'Lê Văn C'],
            'Giới tính': ['Nam', 'Nữ', 'Nam'],
            'Ngày sinh': ['01/01/2000', '15/05/2000', '10/12/1999'],
            'Lớp': ['21DTHD1', '21DTHD1', '21DTHD2'],
            'Điểm 10': [8.5, 7.8, 9.0],
            'Xếp loại': ['Giỏi', 'Khá', 'Xuất sắc'],
            'Ghi chú': ['', '', 'Sinh viên xuất sắc']
        }
        
        df = pd.DataFrame(template_data)
        
        if template_type == 'excel':
            # Create Excel file
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Students')
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name='student_template.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # Create CSV file
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            response = app.response_class(
                output.getvalue(),
                mimetype='text/csv',
                headers={"Content-disposition": "attachment; filename=student_template.csv"}
            )
            return response
            
    except Exception as e:
        logger.error(f"Error creating template: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    # Try to load sample data on startup
    load_sample_data()

    app.run(debug=True, host='0.0.0.0', port=5002)
